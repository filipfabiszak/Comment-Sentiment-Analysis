from bs4 import BeautifulSoup
import urllib.request
import json
import openpyxl
from openpyxl.styles import Font
from commentHelper import *

'''Script to retrieve Kinja comments into an excel file'''

# row to start inputting data into excel spreadsheet
excelRow = 2
wb = openpyxl.load_workbook('KinjaComments.xlsx')
delsheet = wb.get_sheet_by_name("Sheet1")
wb.remove_sheet(delsheet)
wb.create_sheet("Sheet1")
sheet = wb.get_sheet_by_name("Sheet1")
debugCounter = 1

print("This program only works on Kinja websites, links intended for scraping should include a 10 digit article code.")
print("As is currently implemented, the following links in 'KinjaLinks.txt' will be ignored:")

validLinks = getLinks()

sheet.cell(row = 1, column = 1).value = "Kinja Article ID"
sheet.cell(row = 1, column = 2).value = "Article Info"
sheet.cell(row = 1, column = 3).value = "Comments"
sheet.cell(row = 1, column = 4).value = "Commenter Name"
sheet.cell(row = 1, column = 5).value = "Commenter Target Name"
sheet.cell(row = 1, column = 6).value = "Word Count"
sheet.cell(row = 1, column = 7).value = "Character Count"
sheet.cell(row = 1, column = 8).value = "Likes"
sheet.cell(row = 1, column = 9).value = "Images"

for articleLink in validLinks:

    startIndex = 0
    numberOfComments = 0
    approvedChildComments = 0

    currentSource = findSource(articleLink)
    currentCode = findCode(articleLink)
    webURL = currentSource + currentCode
    # print("")
    # print("link: " + articleLink)

    try:
        web = urllib.request.urlopen(webURL)
    except:
        print("Error, cannot open URL: " + webURL)
        continue

    soup = BeautifulSoup(web.read(), "html.parser")

    #Find the specific HTML element that holds the number of total replies
    try:
        headline = findHeadline(soup)
        author = findAuthor(soup)
        totalNumComments = findReplies(soup)
    except:
        print("Error, cannot find headline or totalcomments")

    headlineRow = excelRow
    sheet.cell(row=headlineRow, column=1).value = currentCode
    sheet.cell(row=headlineRow, column=2).hyperlink = webURL
    sheet.cell(row=headlineRow+1, column=2).value = "Article Title: " + headline
    sheet.cell(row=headlineRow+2, column=2).value = "Author: " + author

    # stats to keep track of
    avgMainLikes = 0
    avgMainWord = 0
    avgMainChar = 0
    avgChildLikes = 0
    avgChildWord = 0
    avgChildChar = 0
    imageCount = 0

    # Keep looping until we get all comments (calling different JSON links)
    while startIndex < totalNumComments:

        # Link can be changed to included non approved comments as well
        jsonURL = currentSource + "api/comments/views/replies/{0}?dap=true&startIndex={1}&maxReturned" \
                  "=100&maxChildren=100&approvedOnly=true&cache=true".format(currentCode, startIndex)

        try:
            page = urllib.request.urlopen(jsonURL).read()
        except:
            print("Error, cannot open URL: " + jsonURL)
            break

        pageString = page.decode('utf-8')
        decoded = json.loads(pageString)
        dataSet = decoded["data"]["items"]

        counter = 0
        while counter < len(dataSet) and len(dataSet) != 0:

            mainComment = dataSet[counter]["reply"]["deprecatedFullPlainText"]
            try:
                mainCommentAuthor = dataSet[counter]["reply"]["author"]["displayName"]
            except:
                mainCommentAuthor = "Failed to find comment author"

            try:
                mainCommentTarget = dataSet[counter]["reply"]["replyMeta"]["parentAuthor"]["displayName"]
            except:
                mainCommentTarget = "Failed to find comment target"

            mainLikes = dataSet[counter]["reply"]["likes"]
            avgMainLikes += mainLikes

            try:
                imageColumn = 8
                imageSet = dataSet[counter]["reply"]["images"]
                imageCounter = 0
                imageLink = ""
                while imageCounter < len(imageSet):
                    imageType = imageSet[imageCounter]["format"]
                    imageId = imageSet[imageCounter]["id"]
                    imageLink = "https://i.kinja-img.com/gawker-media/image/upload/"+str(imageId)+"."+str(imageType)
                    sheet.cell(row = excelRow, column = imageColumn).hyperlink = imageLink
                    imageCounter += 1
                    imageColumn += 1
                    imageCount += 1
            except:
                print("no main comment image")


            #making sure the comment is not empty
            if mainComment != "":
                mainCommentWordCount = countWords(mainComment)
                mainCommentCharacterCount = countCharacters(mainComment)
                avgMainWord += mainCommentWordCount
                avgMainChar += mainCommentCharacterCount
                try:
                    if mainComment.strip() != "":
                        sheet.cell(row=excelRow, column=3).value = mainComment
                    else:
                        sheet.cell(row=excelRow, column=3).value = "(main image comment)"
                except:
                    sheet.cell(row = excelRow, column = 3).value = "error string"
                sheet.cell(row = excelRow, column = 3).font = Font(bold = True)
                sheet.cell(row = excelRow, column = 4).value = mainCommentAuthor
                sheet.cell(row = excelRow, column = 5).value = mainCommentTarget
                sheet.cell(row = excelRow, column = 6).value = mainCommentWordCount
                sheet.cell(row = excelRow, column = 7).value = mainCommentCharacterCount
                sheet.cell(row = excelRow, column = 8).value = mainLikes
                excelRow+=1
            elif imageLink != None:
                sheet.cell(row = excelRow, column = 3).value = "(main image comment)"
                sheet.cell(row = excelRow, column = 3).font = Font(bold = True)
                sheet.cell(row = excelRow, column = 4).value = mainCommentAuthor
                sheet.cell(row = excelRow, column = 5).value = mainCommentTarget
                sheet.cell(row = excelRow, column = 6).value = 0
                sheet.cell(row = excelRow, column = 7).value = 0
                sheet.cell(row = excelRow, column = 8).value = mainLikes
                excelRow+=1

            numberOfComments+=1

            childSet = dataSet[counter]["children"]["items"]

            childCounter = 0
            while childCounter < len(childSet):

                childComment = childSet[childCounter]["deprecatedFullPlainText"]

                try:
                    childCommentAuthor = childSet[childCounter]["author"]["displayName"]
                except:
                    childCommentAuthor = "Failed to find comment author"
                try:
                    childCommentTarget = childSet[childCounter]["replyMeta"]["parentAuthor"]["displayName"]
                except:
                    try:
                        childCommentTarget = childSet[childCounter]["replyMeta"]["parentAuthors"][0]["displayName"]
                    except:
                        childCommentTarget = "No Target Specified" # hacky fix for kinja having different replyMeta

                childLikes = childSet[childCounter]["likes"]
                avgChildLikes += childLikes

                try:
                    imageSet = childSet[childCounter]["images"]
                    imageColumn = 8
                    imageCounter = 0
                    imageLink = ""
                    while imageCounter < len(imageSet):
                        imageType = imageSet[imageCounter]["format"]
                        imageId = imageSet[imageCounter]["id"]
                        imageLink += "https://i.kinja-img.com/gawker-media/image/upload/"+str(imageId)+"."+str(imageType) + "   "
                        sheet.cell(row = excelRow, column = imageColumn).hyperlink = str(imageLink)
                        imageCounter += 1
                        imageColumn += 1
                        imageCount += 1
                except:
                    imageLink = "nil"
                    print("no child comment image")

                if childComment != "":
                    childWordLen = countWords(childComment)
                    childCharLen = countCharacters(childComment)
                    avgChildWord += childWordLen
                    avgChildChar += childCharLen

                    try:
                        if childComment.strip() != "":
                            sheet.cell(row=excelRow, column=3).value = childComment
                        else:
                            sheet.cell(row=excelRow, column=3).value = "(child image comment)"
                    except:
                        sheet.cell(row=excelRow, column=3).value = "error string"
                    sheet.cell(row = excelRow, column = 4).value = childCommentAuthor
                    sheet.cell(row = excelRow, column = 5).value = childCommentTarget
                    sheet.cell(row = excelRow, column = 6).value = childWordLen
                    sheet.cell(row = excelRow, column = 7).value = childCharLen
                    sheet.cell(row = excelRow, column = 8).value = childLikes
                    excelRow+=1
                    approvedChildComments+= 1
                elif imageLink != None:
                    sheet.cell(row = excelRow, column = 3).value = "(child image comment)"
                    sheet.cell(row = excelRow, column = 4).value = childCommentAuthor
                    sheet.cell(row = excelRow, column = 5).value = childCommentTarget
                    sheet.cell(row = excelRow, column = 6).value = 0
                    sheet.cell(row = excelRow, column = 7).value = 0
                    sheet.cell(row = excelRow, column = 8).value = childLikes
                    excelRow+=1
                    approvedChildComments+= 1
                childCounter+=1
            counter += 1
        startIndex+=100
        # print("Adding another 100 to get comments " + str(startIndex))

    #code to parse into excel document chosen above
    sheet.cell(row = headlineRow+3, column = 2).value = "Number of total comments: {}".format(str(totalNumComments))
    sheet.cell(row = headlineRow+4, column = 2).value = "Number of total posted comments: {}".\
        format(str(numberOfComments + approvedChildComments))
    sheet.cell(row = headlineRow+5, column = 2).value = "Number of main comments: {}"\
        .format(str(numberOfComments))
    sheet.cell(row = headlineRow+6, column = 2).value = "Number of child comments: {}".\
        format(str(approvedChildComments))

    try:
        sheet.cell(row = headlineRow+7, column = 2).value = "Average Main Comment Likes: {}".\
                    format(str(avgMainLikes/numberOfComments))
        sheet.cell(row = headlineRow+8, column = 2).value = "Average Main Comment Word Count: {}".\
                    format(str(avgMainWord/numberOfComments))
        sheet.cell(row = headlineRow+9, column = 2).value = "Average Main Comment Character Count: {}".\
                    format(str(avgMainChar/numberOfComments))
    except:
        sheet.cell(row = headlineRow+7, column = 2).value = "Average Main Comment Likes: {}".\
                    format("0")
        sheet.cell(row = headlineRow+8, column = 2).value = "Average Main Comment Word Count: {}".\
                    format("0")
        sheet.cell(row = headlineRow+9, column = 2).value = "Average Main Comment Character Count: {}".\
                    format("0")

    try:
        sheet.cell(row = headlineRow+10, column = 2).value = "Average Child Comment Likes: {}".\
                    format(str(avgChildLikes/approvedChildComments))
        sheet.cell(row = headlineRow+11, column = 2).value = "Average Child Comment Word Count: {}".\
                    format(str(avgChildWord/approvedChildComments))
        sheet.cell(row = headlineRow+12, column = 2).value = "Average Child Comment Character Count: {}".\
                    format(str(avgChildChar/approvedChildComments))
    except:
        sheet.cell(row = headlineRow+10, column = 2).value = "Average Child Comment Likes: {}".\
                    format("0")
        sheet.cell(row = headlineRow+11, column = 2).value = "Average Child Comment Word Count: {}".\
                    format("0")
        sheet.cell(row = headlineRow+12, column = 2).value = "Average Child Comment Character Count: {}".\
                    format("0")

    sheet.cell(row = headlineRow+13, column = 2).value = "Number of images: {}".\
            format(str(imageCount))


    print("Article {} done".format(debugCounter))
    debugCounter += 1

    if(numberOfComments < 9):
        excelRow += 15
    else:
        excelRow += 2
    wb.save('KinjaComments.xlsx')

wb.save('KinjaComments.xlsx')
print("Kinja Comment Scraping Finished")
