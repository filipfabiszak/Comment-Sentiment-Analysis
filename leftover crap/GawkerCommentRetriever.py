from bs4 import BeautifulSoup
import urllib.request
import json
import openpyxl
from openpyxl.styles import Font

from commentHelper import countWords
from commentHelper import countCharacters
from commentHelper import getCode2
from commentHelper import findCode
from commentHelper import findHeadline
from commentHelper import findReplies

'''Script to retrieve Jezebel comments into an excel file'''

#row to start parsing at (change if needed)
excelRow = 1
wb = openpyxl.load_workbook('jezebeltest.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")

#Change index here to look for specific articles
articleStartIndex = 0
articleEndIndex = 1

print("Note: This program only works on Jezebel, links entered must include the 10 digit article code.")
print("")
userInput = input("Press Enter for mass scraping OR paste link/article code for specific link scraping: ")
if userInput == "":
    getSpecific = False

    articleStartIndex = int(input("Please choose the start index for the articles you want to scrap: ")) - 1
    articleEndIndex = articleStartIndex + int(input("Please choose how many articles you would like to scrap: "))

else:
    getSpecific = True

debugCounter = 1
articleCodes = getCode2()

for i in range(articleStartIndex, articleEndIndex):
#Index to keep track of the comments (used to change link and get new comments)
    startIndex = 0
    numberOfComments = 0
    approvedChildComments = 0

    currentCode = articleCodes[i]
    webURL = "http://www.jezebel.com/{}".format(currentCode)
    if getSpecific:
        webURL = userInput
        articleCodes[i] = findCode(webURL)

    #Open request to webpage
    try:
        web = urllib.request.urlopen(webURL)
    except:
        print("Error, cannot open URL")
        break

    soup = BeautifulSoup(web.read(), "html.parser")

    #Find the specific HTML element that holds the number of total replies
    try:
        r = findReplies(soup)
        headline = findHeadline(soup)
    except:
        print("Error, cannot find headline (maybe headline does not exist?)")
        headline = "(no headline)"


    headlineRow = excelRow
    sheet.cell(row=headlineRow, column=1).value = (headline)
    #stats to keep track of
    avgMainWord = 0
    avgMainChar = 0
    avgChildWord = 0
    avgChildChar = 0

    imageCount = 0

    #Keep looping until we get all comments (calling different JSON links)
    while startIndex < r:

        #Link can be changed to included non approved comments as well
        jsonURL = "http://jezebel.com/api/comments/views/replies/{0}?dap=true&startIndex={1}&maxReturned" \
                  "=100&maxChildren=100&approvedOnly=true&cache=true".format(articleCodes[i], startIndex)


        page = urllib.request.urlopen(jsonURL).read()
        pageString = page.decode('utf-8')

        #Turns JSON file into dictionary
        decoded = json.loads(pageString)
        dataSet = decoded["data"]["items"]
        counter = 0

        while counter < len(dataSet):
            #Going through the content and taking what we need
            htmlLines = BeautifulSoup(dataSet[counter]["reply"]["display"], "html.parser")
            mainComment = htmlLines.findAll('p')
            try:
                imageColumn = 5
                imageSet = dataSet[counter]["reply"]["images"]
                imageCounter = 0
                imageLink = ""
                while imageCounter < len(imageSet):
                    imageType = BeautifulSoup(imageSet[imageCounter]["format"], "html.parser")
                    imageId = BeautifulSoup(imageSet[imageCounter]["id"], "html.parser")
                    imageLink = "https://i.kinja-img.com/gawker-media/image/upload/"+str(imageId)+"."+str(imageType) + "   "
                    sheet.cell(row = excelRow, column = imageColumn).hyperlink = str(imageLink)
                    imageCounter += 1
                    imageColumn += 1
                    imageCount += 1
            except:
                print("fail")

            fullComments = ""
            if len(mainComment) > 0:
                pass
            else:
                mainComment = htmlLines.findAll('h2')
                if not len(mainComment) > 0:
                    mainComment = htmlLines.findAll('li')



            for comment in mainComment:
                text = comment.getText()
                fullComments += " " + text


            #making sure the comment is not empty
            if mainComment != "":
                mainWordLen = countWords(fullComments)
                mainCharLen = countCharacters(fullComments)
                try:
                    if fullComments.strip() != "":
                        sheet.cell(row=excelRow, column=2).value = fullComments
                    else:
                        sheet.cell(row=excelRow, column=2).value = "(main image comment)"
                except:
                    sheet.cell(row=excelRow, column=2).value = "error string"

                sheet.cell(row = excelRow, column = 3).value = mainWordLen
                avgMainWord += mainWordLen
                sheet.cell(row = excelRow, column =4).value = mainCharLen
                avgMainChar += mainCharLen
                sheet.cell(row=excelRow, column=2).font = Font(bold = True)
                excelRow+=1
            elif imageLink != None:
                sheet.cell(row=excelRow, column=2).value = "(main image comment)"
                sheet.cell(row = excelRow, column = 3).value = 0
                sheet.cell(row = excelRow, column =4).value = 0
                excelRow+=1


            numberOfComments+=1
            childSet = dataSet[counter]["children"]["items"]

            childCounter = 0
            while childCounter < len(childSet):
                try:
                    imageSet = childSet[childCounter]["images"]
                    imageColumn = 5
                    imageCounter = 0
                    imageLink = ""
                    while imageCounter < len(imageSet):
                        imageType = BeautifulSoup(imageSet[imageCounter]["format"], "html.parser")
                        imageId = BeautifulSoup(imageSet[imageCounter]["id"], "html.parser")
                        imageLink += "https://i.kinja-img.com/gawker-media/image/upload/"+str(imageId)+"."+str(imageType) + "   "
                        sheet.cell(row = excelRow, column = imageColumn).hyperlink = str(imageLink)
                        imageCounter += 1
                        imageColumn += 1
                        imageCount += 1

                except:
                    imageLink = "nil"

                htmlLine = BeautifulSoup(childSet[childCounter]["display"], "html.parser")
                childComment = htmlLine.findAll('p')
                fullComment = ""
                if len(childComment) > 0:
                    pass
                else:
                    childComment = htmlLines.findAll('h2')
                    if not len(childComment) > 0:
                        childComment = htmlLines.findAll('li')
                for comment in childComment:
                    text = comment.getText()
                    fullComment += " " + text

                childCounter+=1
                if fullComment != "":
                    childWordLen = countWords(fullComment)
                    childCharLen = countCharacters(fullComment)
                    try:
                        if fullComment.strip() != "":
                            sheet.cell(row=excelRow, column=2).value = fullComment
                        else:
                            sheet.cell(row=excelRow, column=2).value = "(child image comment)"
                    except:
                        sheet.cell(row=excelRow, column=2).value = "error string"
                    sheet.cell(row = excelRow, column = 3).value = childWordLen
                    avgChildWord += childWordLen
                    sheet.cell(row = excelRow, column =4).value = childCharLen
                    avgChildChar += childCharLen
                    excelRow+=1
                    approvedChildComments+= 1
                elif imageLink != None:
                    sheet.cell(row=excelRow, column=2).value = "(child image comment)"
                    sheet.cell(row = excelRow, column = 3).value = 0
                    sheet.cell(row = excelRow, column =4).value = 0
                    excelRow+=1
                    approvedChildComments+= 1
            counter += 1
        startIndex+=100
        print(startIndex)



    #code to parse into excel document chosen above
    headlineRow += 1
    sheet.cell(row=headlineRow, column = 1).value = "Article no: " + str(i + 1)
    sheet.cell(row = headlineRow+1, column = 1).value = "Number of total comments: {}".format(str(r))
    sheet.cell(row = headlineRow+2, column = 1).value = "Number of total posted comments: {}".\
        format(str(numberOfComments + approvedChildComments))
    sheet.cell(row = headlineRow+3, column = 1).value = "Number of main comments: {}"\
        .format(str(numberOfComments))
    sheet.cell(row = headlineRow+4, column = 1).value = "Number of child comments: {}".\
        format(str(approvedChildComments))
    sheet.cell(row = headlineRow+5, column = 1).value = "Average Main Comment Word Count: {}".\
            format(str(avgMainWord/numberOfComments))
    sheet.cell(row = headlineRow+6, column = 1).value = "Average Main Comment Character Count: {}".\
                format(str(avgMainChar/numberOfComments))
    try:
        sheet.cell(row = headlineRow+7, column = 1).value = "Average Child Comment Word Count: {}".\
                    format(str(avgChildWord/approvedChildComments))
        sheet.cell(row = headlineRow+8, column = 1).value = "Average Child Comment Character Count: {}".\
                format(str(avgChildChar/approvedChildComments))
    except:
        sheet.cell(row = headlineRow+7, column = 1).value = "Average Child Comment Word Count: {}".\
                    format("0")
        sheet.cell(row = headlineRow+8, column = 1).value = "Average Child Comment Character Count: {}".\
                format("0")

    sheet.cell(row = headlineRow+9, column = 1).value = "Number of images: {}".\
            format(str(imageCount))


    print("{} Article done".format(debugCounter))
    debugCounter+=1
    if(numberOfComments < 9):
        excelRow += 11
    else:
        excelRow += 2
    wb.save('jezebeltest.xlsx')

wb.save('jezebeltest.xlsx')
