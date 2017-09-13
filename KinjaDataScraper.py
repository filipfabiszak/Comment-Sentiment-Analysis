from bs4 import BeautifulSoup
import urllib.request
import json
import openpyxl
from commentHelper import *

'''Script to retrieve comment data from Kinja articles'''

# row to start parsing at (change if needed)
excelRow = 2
wb = openpyxl.load_workbook('KinjaData.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")
debugCounter = 1
approved = True

print("This program only works on Kinja websites, links intended for scraping should include a 10 digit article code.")
print("As is currently implemented, the following links in 'KinjaLinks.txt' will be ignored:")

validLinks = getLinks()

sheet.cell(row = 1, column = 1).value = "Article Link"
sheet.cell(row = 1, column = 2).value = "No. of Total Comments"
sheet.cell(row = 1, column = 3).value = "No. of Posted Comments"
sheet.cell(row = 1, column = 4).value = "No. of Main"
sheet.cell(row = 1, column = 5).value = "No. of Children"
sheet.cell(row = 1, column = 6).value = "Avg Main Likes"
sheet.cell(row = 1, column = 7).value = "Avg Main Word Count"
sheet.cell(row = 1, column = 8).value = "Avg Main Character Count"
sheet.cell(row = 1, column = 9).value = "Avg Child Likes"
sheet.cell(row = 1, column = 10).value = "Avg Child Word Count"
sheet.cell(row = 1, column = 11).value = "Avg Child Character Count"
sheet.cell(row = 1, column = 12).value = "Number of images"

for articleLink in validLinks:

    startIndex = 0
    numberOfComments = 0
    approvedChildComments = 0

    currentSource = findSource(articleLink)
    currentCode = findCode(articleLink)
    webURL = currentSource + currentCode

    try:
        web = urllib.request.urlopen(webURL)
    except:
        print("Error, cannot open URL: " + webURL)
        continue

    soup = BeautifulSoup(web.read(), "html.parser")

    # Find the specific HTML element that holds the number of total replies
    try:
        totalNumComments = findReplies(soup)
    except:
        print("Error, cannot find replies")
        headline = "(no headline)"


    headlineRow = excelRow
    # sheet.cell(row=headlineRow, column=1).value = (headline)
    # stats to keep track of
    avgMainLikes = 0
    avgMainWord = 0
    avgMainChar = 0
    avgChildLikes = 0
    avgChildWord = 0
    avgChildChar = 0
    imageCount = 0

    # Keep looping until we get all comments (calling different JSON links)
    while startIndex < totalNumComments:

        if approved:
            jsonURL = currentSource + "api/comments/views/replies/{0}?dap=true&startIndex={1}&maxReturned" \
                  "=100&maxChildren=100&approvedOnly=true&cache=true".format(currentCode, startIndex)
        else:
            jsonURL = currentSource + "api/comments/views/replies/{0}?dap=true&startIndex={1}&maxReturned" \
                  "=100&maxChildren=100&approvedOnly=false&cache=true".format(currentCode, startIndex)

        try:
            page = urllib.request.urlopen(jsonURL).read()
        except:
            print("Error, cannot open URL: " + jsonURL)
            break

        pageString = page.decode('utf-8')
        decoded = json.loads(pageString)
        dataSet = decoded["data"]["items"]

        counter = 0
        while counter < len(dataSet) and len(dataSet) != 0:

            mainComment = dataSet[counter]["reply"]["deprecatedFullPlainText"]
            try:
                imageCount += len(dataSet[counter]["reply"]["images"])
            except:
                pass
            avgMainLikes += dataSet[counter]["reply"]["likes"]

            if mainComment != "":
                mainCommentWordCount = countWords(mainComment)
                mainCommentCharacterCount = countCharacters(mainComment)
                avgMainWord += mainCommentWordCount
                avgMainChar += mainCommentCharacterCount

            numberOfComments+=1

            childSet = dataSet[counter]["children"]["items"]

            childCounter = 0
            while childCounter < len(childSet):

                childComment = childSet[childCounter]["deprecatedFullPlainText"]
                try:
                    imageCount += len(childSet[childCounter]["images"])
                except:
                    pass
                avgChildLikes += childSet[childCounter]["likes"]

                if childComment != "":
                    childWordLen = countWords(childComment)
                    childCharLen = countCharacters(childComment)
                    avgChildWord += childWordLen
                    avgChildChar += childCharLen
                approvedChildComments+= 1
                childCounter+=1
            counter += 1
        startIndex+=100

    sheet.cell(row = excelRow, column = 1).hyperlink = webURL
    sheet.cell(row = excelRow, column = 2).value = totalNumComments
    sheet.cell(row = excelRow, column = 3).value = numberOfComments + approvedChildComments
    sheet.cell(row = excelRow, column = 4).value = numberOfComments
    sheet.cell(row = excelRow, column = 5).value = approvedChildComments
    try:
        sheet.cell(row = excelRow, column = 6).value = avgMainLikes/numberOfComments
        sheet.cell(row = excelRow, column = 7).value = avgMainWord/numberOfComments
        sheet.cell(row = excelRow, column = 8).value = avgMainChar/numberOfComments
    except:
        sheet.cell(row = excelRow, column = 6).value = numberOfComments
        sheet.cell(row = excelRow, column = 7).value = numberOfComments
        sheet.cell(row = excelRow, column = 8).value = numberOfComments
    try:
        sheet.cell(row = excelRow, column = 9).value = avgChildLikes/approvedChildComments
        sheet.cell(row = excelRow, column = 10).value = avgChildWord/approvedChildComments
        sheet.cell(row = excelRow, column = 11).value = avgChildChar/approvedChildComments
    except:
        sheet.cell(row = excelRow, column = 9).value = approvedChildComments
        sheet.cell(row = excelRow, column = 10).value = approvedChildComments
        sheet.cell(row = excelRow, column = 11).value = approvedChildComments

    sheet.cell(row = excelRow, column = 12).value = imageCount

    # if approved:
    # else:
    #     sheet.cell(row = excelRow, column = 3).value = (numberOfComments + approvedChildComments)
    #     sheet.cell(row = excelRow, column = 6).value = ((numberOfComments))
    #     sheet.cell(row = excelRow, column = 8).value = ((approvedChildComments))
    #     sheet.cell(row = excelRow, column = 13).value = ((avgMainWord/numberOfComments))
    #     sheet.cell(row = excelRow, column = 15).value = ((avgMainChar/numberOfComments))
    #     if approvedChildComments != 0:
    #         sheet.cell(row = excelRow, column = 14).value = ((avgChildWord/approvedChildComments))
    #         sheet.cell(row = excelRow, column = 16).value = ((avgChildChar/approvedChildComments))
    #     else:
    #         sheet.cell(row = excelRow, column = 14).value = ((approvedChildComments))
    #         sheet.cell(row = excelRow, column = 16).value = ((approvedChildComments))

    print("Article {} done".format(debugCounter))
    debugCounter+=1
    excelRow += 1
    wb.save('KinjaData.xlsx')

wb.save('KinjaData.xlsx')
