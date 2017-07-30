from bs4 import BeautifulSoup
import urllib.request
import json
import openpyxl
from openpyxl.styles import Font

from commentHelper import countWords
from commentHelper import countCharacters
from commentHelper import getCode2
from commentHelper import findCode
from commentHelper import findHeadline
from commentHelper import findReplies

'''Script to retrieve Jezebel comments into an excel file'''

#row to start parsing at (change if needed)
excelRow = 2
#Workbook and worksheet
wb = openpyxl.load_workbook('GawkerScrape.xlsx')
sheet = wb.get_sheet_by_name("JezebelScrape2")

#Change index here to look for specific articles
articleStartIndex = 0
articleEndIndex = 1

print("Note: This program only works on Jezebel, links entered must include the 10 digit article code.")
print("")
userInput = input("Press Enter for mass scraping OR paste link/article code for specific link scraping: ")
if userInput == "":
    getSpecific = False

    articleStartIndex = int(input("Please choose the start index for the articles you want to scrap: ")) - 1
    articleEndIndex = articleStartIndex + int(input("Please choose how many articles you would like to scrap: "))
    
else:
    excelRow = int(input("Please indicate the excel row you want the information to be parsed into"))
    getSpecific = True

debugCounter = 1
#get jezebel article codes
articleCodes = getCode2()

for i in range(articleStartIndex, articleEndIndex):
#Index to keep track of the comments (used to change link and get new comments)
    startIndex = 0
    numberOfComments = 0
    approvedChildComments = 0

    currentCode = articleCodes[i]
    webURL = "http://jezebel.com/{}".format(currentCode)
    if getSpecific:
        webURL = userInput
        currentCode = findCode(webURL)

    #Open request to webpage
    try:
        web = urllib.request.urlopen(webURL)
    except:
        print("Error, cannot open URL")
        break

    soup = BeautifulSoup(web.read(), "html.parser")

    #Find the specific HTML element that holds the number of total replies
    try:
        r = findReplies(soup)
        headline = findHeadline(soup)
    except:
        print("Error, cannot find proper HTML elements (Maybe the link you entered is wrong)")
        break


    headlineRow = excelRow
    sheet.cell(row=headlineRow, column=1).value = (headline)
    #stats to keep track of
    avgMainWord = 0
    avgMainChar = 0
    avgChildWord = 0
    avgChildChar = 0

    #Keep looping until we get all comments (calling different JSON links)
    while startIndex < r:



        #Link can be changed to included non approved comments as well
        #Change approvecChildrenOnly and approvedStartersOnly to false

        jsonJezbelURL = "http://jezebel.com/api/core/reply/{0}/replies?currentBlogId=39&startIndex={1}&maxReturned=" \
                        "100&withLikeCounts=true&maxChildren=4&approvedChildrenOnly=true&approvedStartersOnly=true&cach" \
                        "e=true".format(currentCode, startIndex)

        page = urllib.request.urlopen(jsonJezbelURL).read()
        pageString = page.decode('utf-8')

        #Turns JSON file into dictionary
        decoded = json.loads(pageString)
        dataSet = decoded["data"]["items"]

        counter = 0

        while counter < len(dataSet):
            #Going through the content and taking what we need
            htmlLines = BeautifulSoup(dataSet[counter]["reply"]["display"], "html.parser")
            mainComment = htmlLines.findAll('p')
            fullComments = ""
            #concatenate all parts of the comments (from the JSON data list) into a full string
            for comment in mainComment:
                text = comment.getText()
                fullComments += " " + text
            #making sure the comment is not empty
            if mainComment != "":
                #parsing main comments and data into file
                mainWordLen = countWords(fullComments)
                mainCharLen = countCharacters(fullComments)
                sheet.cell(row=excelRow, column=2).value = fullComments
                sheet.cell(row = excelRow, column = 3).value = mainWordLen
                avgMainWord += mainWordLen
                sheet.cell(row = excelRow, column =4).value = mainCharLen
                avgMainChar += mainCharLen
                sheet.cell(row=excelRow, column=2).font = Font(bold = True)
                excelRow+=1

            childSet = dataSet[counter]["children"]["items"]
            numberOfComments+=1

            childCounter = 0
            while childCounter < len(childSet):
                htmlLine = BeautifulSoup(childSet[childCounter]["display"], "html.parser")
                childComment = htmlLine.findAll('p')
                fullComment = ""
                for comment in childComment:
                    text = comment.getText()
                    fullComment += " " + text

                childCounter+=1
                if fullComment != "":
                    #parsing child comments and data into file
                    childWordLen = countWords(fullComment)
                    childCharLen = countCharacters(fullComment)
                    sheet.cell(row=excelRow, column=2).value = fullComment
                    sheet.cell(row = excelRow, column = 3).value = childWordLen
                    avgChildWord += childWordLen
                    sheet.cell(row = excelRow, column =4).value = childCharLen
                    avgChildChar += childCharLen
                    excelRow+=1
                    approvedChildComments+= 1
            counter += 1
        startIndex+=100

    #Code to parse into chosen excel file above
    sheet.cell(row = headlineRow+1, column = 1).value = "Number of total comments: {}".format(str(r))
    sheet.cell(row = headlineRow+2, column = 1).value = "Number of total approved comments: {}".\
        format(str(numberOfComments + approvedChildComments))
    sheet.cell(row = headlineRow+3, column = 1).value = "Number of approved main comments: {}"\
        .format(str(numberOfComments))
    sheet.cell(row = headlineRow+4, column = 1).value = "Number of approved child comments: {}".\
        format(str(approvedChildComments))
    sheet.cell(row = headlineRow+5, column = 1).value = "Average Main Comment Word Count: {}".\
            format(str(avgMainWord/numberOfComments))
    sheet.cell(row = headlineRow+6, column = 1).value = "Average Main Comment Character Count: {}".\
                format(str(avgMainChar/numberOfComments))
    sheet.cell(row = headlineRow+7, column = 1).value = "Average Child Comment Word Count: {}".\
                format(str(avgChildWord/approvedChildComments))
    sheet.cell(row = headlineRow+8, column = 1).value = "Average Child Comment Character Count: {}".\
            format(str(avgChildChar/approvedChildComments))


    print("{} Article done".format(debugCounter))
    debugCounter+=1
    excelRow += 2

wb.save('GawkerScrape.xlsx')