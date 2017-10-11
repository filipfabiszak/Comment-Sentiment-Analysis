import openpyxl
from commentHelper import *

'''Script to scrape Kinja articles completely'''

excelRow = 2
debugcounter = 0
wb = openpyxl.load_workbook('KinjaArticles.xlsx')
delsheet = wb.get_sheet_by_name("Sheet1")
wb.remove_sheet(delsheet)
wb.create_sheet("Sheet1")
sheet = wb.get_sheet_by_name("Sheet1")
debugCounter = 1
approved = True

print("This program only works on Kinja websites, links intended for scraping should include a 10 digit article code.")
print("As is currently implemented, the following links in 'KinjaLinks.txt' will be ignored:")

validLinks = getLinks()

sheet.cell(row = 1, column = 1).value = "Article Link"
sheet.cell(row = 1, column = 2).value = "Title"
sheet.cell(row = 1, column = 3).value = "Author"
sheet.cell(row = 1, column = 4).value = "Date Posted"

sheet.cell(row = 1, column = 5).value = "Num Comments"
sheet.cell(row = 1, column = 6).value = "Num Likes"
sheet.cell(row = 1, column = 7).value = "Word Count"
sheet.cell(row = 1, column = 8).value = "Character Count"
sheet.cell(row = 1, column = 9).value = "Article Text"

for articleLink in validLinks:

    currentSource = findSource(articleLink)
    currentCode = findCode(articleLink)
    webURL = currentSource + currentCode
    try:
        currentArticle = getArticle(webURL)
    except:
        continue

    web = urllib.request.urlopen(webURL)
    soup = BeautifulSoup(web.read(), "html.parser")

    articleCharCount = countCharacters(currentArticle)
    sheet.cell(row = excelRow, column = 1).hyperlink = webURL
    sheet.cell(row = excelRow, column = 2).value = findHeadline(soup)
    sheet.cell(row = excelRow, column = 3).value = findAuthor(soup)
    sheet.cell(row = excelRow, column = 4).value = findDate(soup)
    sheet.cell(row = excelRow, column = 5).value = findReplies(soup)
    sheet.cell(row = excelRow, column = 6).value = findLikes(soup)
    sheet.cell(row = excelRow, column = 7).value = countWords(currentArticle)
    sheet.cell(row = excelRow, column = 8).value = articleCharCount
    if articleCharCount > 32767:
        sheet.cell(row = excelRow, column = 9).value = currentArticle[:32767]
        excelRow += 1
        sheet.cell(row = excelRow, column = 9).value = currentArticle[32767:]
    else:
        sheet.cell(row = excelRow, column = 9).value = currentArticle

    excelRow += 1
    print("Article {} done".format(debugCounter))
    debugCounter += 1
    wb.save("KinjaArticles.xlsx")

wb.save("KinjaArticles.xlsx")
print("Kinja Article Scraping Finished")
